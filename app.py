import streamlit as st
import pdfplumber
import pandas as pd
import re
import json
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import List, Dict, Any, Union
import zipfile
import tempfile
import os

# Set page config
st.set_page_config(
    page_title="PDF/JSON to Excel Converter",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

def detect_input_type_from_content(content: bytes, filename: str) -> str:
    """
    Detect if the input is PDF or JSON based on content and filename
    """
    file_extension = Path(filename).suffix.lower()
    if file_extension == '.pdf':
        return 'pdf'
    elif file_extension == '.json':
        return 'json'
    else:
        # Try to determine by content
        try:
            text_content = content.decode('utf-8').strip()
            if text_content.startswith('{') or text_content.startswith('['):
                return 'json'
        except:
            pass
        return 'pdf'  # Default to PDF

def extract_from_json_content(json_content: str) -> List[Dict[str, Any]]:
    """
    Extract billing data from JSON content - Updated to handle nested structure
    """
    try:
        data = json.loads(json_content)
    except json.JSONDecodeError as e:
        st.error(f"Error parsing JSON: {e}")
        return []
    
    # Extract lines from various possible locations
    lines = []
    
    # Try multiple possible paths for the billing lines
    possible_paths = [
        # Original structure
        ['upcomingBills', 'lines'],
        ['lines'],
        # Nested structure (like your example)
        ['QuoteDetails', 'upcomingBills', 'lines'],
        # Other possible structures
        ['quote', 'upcomingBills', 'lines'],
        ['data', 'upcomingBills', 'lines'],
    ]
    
    for path in possible_paths:
        try:
            current_data = data
            for key in path:
                if isinstance(current_data, dict) and key in current_data:
                    current_data = current_data[key]
                else:
                    break
            else:
                # If we made it through the whole path, we found the lines
                if isinstance(current_data, list) and current_data:
                    lines = current_data
                    st.success(f"Found billing lines at: {' → '.join(path)}")
                    break
        except (KeyError, TypeError):
            continue
    
    if not lines:
        st.warning("No billing lines found in JSON. Checking available structure...")
        # Debug: Show the top-level structure
        if isinstance(data, dict):
            st.info(f"Top-level keys found: {list(data.keys())}")
            
            # Look for any nested upcomingBills
            for key, value in data.items():
                if isinstance(value, dict) and 'upcomingBills' in value:
                    st.info(f"Found 'upcomingBills' under key: '{key}'")
                    if 'lines' in value['upcomingBills']:
                        st.info(f"Found 'lines' under '{key}' → upcomingBills")
        return []
    
    st.info(f"Processing {len(lines)} billing lines from JSON")
    
    # Group lines by product description (base product name)
    product_groups = {}
    
    for line in lines:
        description = line.get('description', '')
        total = line.get('total', 0)
        subTotal = line.get('subTotal', 0)  # Also check subTotal
        margins = line.get('margins', [])
        
        # Use subTotal if total is 0, convert from cents to dollars
        amount = total / 100.0 if total != 0 else (subTotal / 100.0 if subTotal != 0 else 0.0)
        
        # Skip zero amounts unless it's the only amount for a product
        if amount == 0:
            # Check if this is a zero line that should be included
            is_credit_line = line.get('isCreditLine', False)
            if not is_credit_line:
                continue
        
        # Extract base product name (remove pricing details in parentheses)
        base_product = re.sub(r'\s*\([^)]*\)\s*$', '', description).strip()
        if not base_product:
            base_product = description
        
        # Check for discounts in margins
        has_discount = 'N'
        for margin in margins:
            if margin.get('percent', 0) > 0 or margin.get('amount', 0) != 0:
                has_discount = 'Y'
                break
        
        # Group by base product name
        if base_product not in product_groups:
            product_groups[base_product] = {
                'product': description,  # Use full description for first item
                'amounts': [],
                'discount': has_discount
            }
        
        # Add amount to the group
        product_groups[base_product]['amounts'].append(amount)
        
        # Update discount status if any item in group has discount
        if has_discount == 'Y':
            product_groups[base_product]['discount'] = 'Y'
    
    if not product_groups:
        st.warning("No valid product groups found after processing billing lines")
        return []
    
    # Convert to table data format
    table_data = []
    for i, (base_product, group_data) in enumerate(product_groups.items(), 1):
        # First item with full product description
        table_data.append({
            'sno': str(i),
            'product': group_data['product'],
            'amount_excl_tax': group_data['amounts'][0] if group_data['amounts'] else 0.0,
            'discount': group_data['discount']
        })
        
        # Additional amounts for the same product (sub-items)
        for amount in group_data['amounts'][1:]:
            table_data.append({
                'sno': '',  # Empty S.no for sub-items
                'product': '',
                'amount_excl_tax': amount,
                'discount': 'N'
            })
    
    st.success(f"Successfully extracted {len(table_data)} items from {len(product_groups)} product groups")
    return table_data

def extract_tables_from_pdf_content(pdf_content: bytes) -> List[List[Dict[str, Any]]]:
    """
    Extract tables from PDF content and return structured data
    """
    all_tables = []
    
    try:
        with pdfplumber.open(io.BytesIO(pdf_content)) as pdf:
            for page in pdf.pages:
                # Extract tables from the page
                tables = page.extract_tables()
                
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                        
                    # Get headers from first row
                    headers = [str(cell).strip() if cell else '' for cell in table[0]]
                    
                    # Find relevant column indices
                    sno_idx = None
                    product_idx = None
                    amount_excl_tax_idx = None
                    discount_idx = None
                    
                    for i, header in enumerate(headers):
                        header_lower = header.lower()
                        if 's.no' in header_lower or 'sno' in header_lower:
                            sno_idx = i
                        elif 'product' in header_lower:
                            product_idx = i
                        elif 'amount' in header_lower and 'excl' in header_lower:
                            amount_excl_tax_idx = i
                        elif 'discount' in header_lower:
                            discount_idx = i
                    
                    # Skip if we don't have essential columns
                    if sno_idx is None or product_idx is None or amount_excl_tax_idx is None:
                        continue
                    
                    # Process data rows
                    table_data = []
                    for row in table[1:]:  # Skip header row
                        if len(row) <= max(sno_idx, product_idx, amount_excl_tax_idx):
                            continue
                            
                        sno = str(row[sno_idx]).strip() if row[sno_idx] else ''
                        product = str(row[product_idx]).strip() if row[product_idx] else ''
                        amount_str = str(row[amount_excl_tax_idx]).strip() if row[amount_excl_tax_idx] else ''
                        discount_str = str(row[discount_idx]).strip() if discount_idx is not None and row[discount_idx] else ''
                        
                        # Skip empty rows
                        if not sno and not product and not amount_str:
                            continue
                        
                        # Parse amount (handle USD prefix, commas, negative values)
                        amount = parse_amount(amount_str)
                        
                        # Determine if there's a discount
                        has_discount = 'Y' if discount_str and discount_str != '' and discount_str != '0' else 'N'
                        
                        table_data.append({
                            'sno': sno,
                            'product': product,
                            'amount_excl_tax': amount,
                            'discount': has_discount
                        })
                    
                    if table_data:
                        all_tables.append(table_data)
    except Exception as e:
        st.error(f"Error processing PDF: {e}")
        return []
    
    return all_tables

def parse_amount(amount_str: str) -> float:
    """
    Parse amount string to float, handling USD prefix, commas, negatives
    """
    if not amount_str or amount_str == '':
        return 0.0
    
    # Remove USD prefix, whitespace, and commas
    cleaned = re.sub(r'USD\s*', '', amount_str, flags=re.IGNORECASE)
    cleaned = cleaned.replace(',', '').strip()
    
    # Handle negative values
    is_negative = cleaned.startswith('-')
    if is_negative:
        cleaned = cleaned[1:]
    
    try:
        value = float(cleaned)
        return -value if is_negative else value
    except ValueError:
        return 0.0

def format_product_text(product_text: str) -> str:
    """
    Format product text with line breaks before Entitlement Number and Billing period
    """
    if not product_text:
        return product_text
    
    # Add line breaks before "Entitlement Number:" and "Billing period:"
    formatted_text = re.sub(r'(Entitlement Number:)', r'\n\1', product_text)
    formatted_text = re.sub(r'(Billing period:)', r'\n\1', formatted_text)
    
    return formatted_text

def group_by_sno(table_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Group items by S.no, handling cases where multiple items share the same S.no
    """
    grouped = {}
    current_sno = None
    
    for item in table_data:
        sno = item['sno']
        
        # If S.no is empty, it belongs to the previous S.no
        if not sno and current_sno:
            sno = current_sno
        else:
            current_sno = sno
        
        if sno not in grouped:
            grouped[sno] = []
        
        grouped[sno].append(item)
    
    return grouped

def create_excel_content(grouped_data: Dict[str, List[Dict[str, Any]]]) -> bytes:
    """
    Create Excel file content with enhanced formatting and formulas in column E
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    
    # Set headers
    ws['A1'] = 'S.no'
    ws['B1'] = 'Discount?'
    ws['C1'] = 'Product'
    ws['D1'] = 'Amount excl. tax'
    ws['E1'] = 'Total'
    
    # Format header row
    header_font = Font(bold=True)
    for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws[col].font = header_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60  # Much wider for product descriptions
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # Create styles
    red_font = Font(color="FF0000")  # Red color for totals
    wrap_text = Alignment(wrap_text=True, vertical='top')
    currency_format = '#,##0.00'
    
    current_row = 2
    formula_rows = []  # Track rows with formulas for grand total
    
    # Sort by S.no for consistent output
    for sno in sorted(grouped_data.keys(), key=lambda x: int(x) if x.isdigit() else 999):
        items = grouped_data[sno]
        
        if not items:
            continue
        
        # First item in the group
        first_item = items[0]
        
        # Set values
        ws[f'A{current_row}'] = sno
        ws[f'B{current_row}'] = first_item['discount']
        
        # Format product text with line breaks
        formatted_product = format_product_text(first_item['product'])
        ws[f'C{current_row}'] = formatted_product
        ws[f'C{current_row}'].alignment = wrap_text
        
        # Format amount with currency
        ws[f'D{current_row}'] = first_item['amount_excl_tax']
        ws[f'D{current_row}'].number_format = currency_format
        
        # Create formula for group total
        group_start_row = current_row
        group_end_row = current_row + len(items) - 1
        
        if len(items) == 1:
            # Single item - formula just references the single cell
            formula = f"=D{current_row}"
        else:
            # Multiple items - sum the range
            formula = f"=SUM(D{group_start_row}:D{group_end_row})"
        
        ws[f'E{current_row}'] = formula
        ws[f'E{current_row}'].number_format = currency_format
        ws[f'E{current_row}'].font = red_font  # Make total red
        
        # Track this row for grand total formula
        formula_rows.append(current_row)
        
        # Set row height to accommodate wrapped text
        ws.row_dimensions[current_row].height = None  # Auto-fit
        
        current_row += 1
        
        # Additional items under the same S.no (sub-items)
        for item in items[1:]:
            ws[f'D{current_row}'] = item['amount_excl_tax']
            ws[f'D{current_row}'].number_format = currency_format
            current_row += 1
    
    # Add grand total with formula
    ws[f'D{current_row}'] = "TOTAL"
    ws[f'D{current_row}'].font = Font(bold=True)
    
    # Create grand total formula that sums all the group totals
    if formula_rows:
        formula_cells = [f"E{row}" for row in formula_rows]
        grand_total_formula = f"=SUM({','.join(formula_cells)})"
        ws[f'E{current_row}'] = grand_total_formula
    else:
        ws[f'E{current_row}'] = 0
    
    ws[f'E{current_row}'].number_format = currency_format
    ws[f'E{current_row}'].font = Font(color="FF0000", bold=True)  # Red and bold for grand total
    
    # Auto-fit row heights for better text display
    for row in range(2, current_row + 1):
        ws.row_dimensions[row].height = None
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

def process_single_file(file_content: bytes, filename: str) -> tuple[bytes, str, str]:
    """
    Process a single file and return Excel content, status, and message
    """
    try:
        # Detect input type
        input_type = detect_input_type_from_content(file_content, filename)
        
        # Extract data based on input type
        if input_type == 'json':
            json_content = file_content.decode('utf-8')
            table_data = extract_from_json_content(json_content)
            if not table_data:
                return None, "error", "No data extracted from JSON"
            combined_data = table_data
        else:  # PDF
            all_tables = extract_tables_from_pdf_content(file_content)
            if not all_tables:
                return None, "error", "No tables found in PDF"
            # Combine all tables
            combined_data = []
            for table in all_tables:
                combined_data.extend(table)
        
        # Group by S.no
        grouped_data = group_by_sno(combined_data)
        
        # Create Excel content
        excel_content = create_excel_content(grouped_data)
        
        return excel_content, "success", f"Successfully processed {len(combined_data)} items into {len(grouped_data)} groups"
        
    except Exception as e:
        return None, "error", f"Error processing file: {str(e)}"

# Main Streamlit App
def main():
    st.title("📊 PDF/JSON to Excel Converter")
    st.markdown("Convert your billing PDFs and JSON files to formatted Excel spreadsheets")
    
    # Sidebar
    with st.sidebar:
        st.header("ℹ️ About")
        st.markdown("""
        This tool converts:
        - **PDF invoices/bills** with tables
        - **JSON billing data** (like Atlassian quotes)
        
        Into formatted Excel files with:
        - Proper grouping by S.no
        - Automatic formulas for totals
        - Currency formatting
        - Red totals column
        - Discount detection
        """)
        
        st.header("📝 Instructions")
        st.markdown("""
        1. Upload one or more files
        2. Wait for processing to complete
        3. Download individual Excel files
        4. Or download all as a ZIP file
        """)
    
    # File uploader
    uploaded_files = st.file_uploader(
        "Choose PDF or JSON files",
        type=['pdf', 'json'],
        accept_multiple_files=True,
        help="Upload one or more PDF or JSON files to convert"
    )
    
    if uploaded_files:
        st.subheader(f"📁 Processing {len(uploaded_files)} file(s)")
        
        # Process files
        processed_files = []
        
        # Create progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for i, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Processing {uploaded_file.name}...")
            
            # Read file content
            file_content = uploaded_file.read()
            
            # Process the file
            excel_content, status, message = process_single_file(file_content, uploaded_file.name)
            
            processed_files.append({
                'filename': uploaded_file.name,
                'excel_content': excel_content,
                'status': status,
                'message': message
            })
            
            # Update progress
            progress_bar.progress((i + 1) / len(uploaded_files))
        
        status_text.text("Processing complete!")
        
        # Display results
        st.subheader("📋 Processing Results")
        
        success_files = []
        error_files = []
        
        for result in processed_files:
            if result['status'] == 'success':
                success_files.append(result)
                st.success(f"✅ {result['filename']}: {result['message']}")
            else:
                error_files.append(result)
                st.error(f"❌ {result['filename']}: {result['message']}")
        
        # Download section
        if success_files:
            st.subheader("📥 Download Files")
            
            # Create two columns for individual downloads and bulk download
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.markdown("**Individual Downloads:**")
                for result in success_files:
                    original_name = Path(result['filename']).stem
                    excel_filename = f"{original_name}_converted.xlsx"
                    
                    st.download_button(
                        label=f"📊 Download {excel_filename}",
                        data=result['excel_content'],
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"download_{original_name}"
                    )
            
            with col2:
                if len(success_files) > 1:
                    st.markdown("**Bulk Download:**")
                    
                    # Create ZIP file
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for result in success_files:
                            original_name = Path(result['filename']).stem
                            excel_filename = f"{original_name}_converted.xlsx"
                            zip_file.writestr(excel_filename, result['excel_content'])
                    
                    zip_buffer.seek(0)
                    
                    st.download_button(
                        label="🗂️ Download All as ZIP",
                        data=zip_buffer.getvalue(),
                        file_name="converted_files.zip",
                        mime="application/zip"
                    )
        
        # Summary statistics
        if success_files or error_files:
            st.subheader("📊 Summary")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Files", len(uploaded_files))
            
            with col2:
                st.metric("Successful", len(success_files), delta=len(success_files))
            
            with col3:
                st.metric("Errors", len(error_files), delta=-len(error_files) if error_files else 0)
    
    else:
        st.info("👆 Upload PDF or JSON files to get started!")
        
        # Show example
        with st.expander("📖 See Example"):
            st.markdown("""
            **Example PDF table structure:**
            ```
            S.no | Product                    | Amount excl. tax | Discount
            1    | Confluence Premium         | 142,500.00      | USD -0.26
            2    | Jira Diagrams             | 9,200.00        | 
                 |                           | -2,073.67       |
            ```
            
            **Example JSON structure:**
            ```json
            {
              "QuoteDetails": {
                "upcomingBills": {
                  "lines": [
                    {
                      "description": "Confluence, Annual, Enterprise",
                      "total": 14250000,
                      "margins": [{"percent": 20}]
                    }
                  ]
                }
              }
            }
            ```
            """)

if __name__ == "__main__":
    main()