import pdfplumber
import pandas as pd
import re
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from typing import List, Dict, Any, Union

def detect_input_type(file_path: str) -> str:
    """
    Detect if the input file is PDF or JSON
    """
    file_extension = Path(file_path).suffix.lower()
    if file_extension == '.pdf':
        return 'pdf'
    elif file_extension == '.json':
        return 'json'
    else:
        # Try to determine by content
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if content.startswith('{') or content.startswith('['):
                    return 'json'
        except:
            pass
        return 'pdf'  # Default to PDF

def extract_from_json(json_path: str) -> List[Dict[str, Any]]:
    """
    Extract billing data from JSON format
    """
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Extract lines from upcomingBills
    lines = []
    if 'upcomingBills' in data and 'lines' in data['upcomingBills']:
        lines = data['upcomingBills']['lines']
    elif 'lines' in data:  # In case the structure is different
        lines = data['lines']
    
    if not lines:
        print("No billing lines found in JSON")
        return []
    
    # Group lines by product description (base product name)
    product_groups = {}
    
    for line in lines:
        description = line.get('description', '')
        total = line.get('total', 0)
        margins = line.get('margins', [])
        
        # Convert total from cents to dollars (assuming it's in cents)
        amount = total / 100.0 if total != 0 else 0.0
        
        # Extract base product name (remove pricing details in parentheses)
        base_product = re.sub(r'\s*\([^)]*\)\s*$', '', description).strip()
        if not base_product:
            base_product = description
        
        # Check for discounts in margins
        has_discount = 'N'
        for margin in margins:
            if margin.get('percent', 0) > 0 or margin.get('amount', 0) != 0:
                has_discount = 'Y'
                break
        
        # Group by base product name
        if base_product not in product_groups:
            product_groups[base_product] = {
                'product': description,  # Use full description for first item
                'amounts': [],
                'discount': has_discount
            }
        
        # Add amount to the group (including zero and negative amounts)
        if amount != 0 or len(product_groups[base_product]['amounts']) == 0:
            product_groups[base_product]['amounts'].append(amount)
        
        # Update discount status if any item in group has discount
        if has_discount == 'Y':
            product_groups[base_product]['discount'] = 'Y'
    
    # Convert to table data format
    table_data = []
    for i, (base_product, group_data) in enumerate(product_groups.items(), 1):
        # First item with full product description
        table_data.append({
            'sno': str(i),
            'product': group_data['product'],
            'amount_excl_tax': group_data['amounts'][0] if group_data['amounts'] else 0.0,
            'discount': group_data['discount']
        })
        
        # Additional amounts for the same product (sub-items)
        for amount in group_data['amounts'][1:]:
            table_data.append({
                'sno': '',  # Empty S.no for sub-items
                'product': '',
                'amount_excl_tax': amount,
                'discount': 'N'
            })
    
    return table_data

def extract_tables_from_pdf(pdf_path: str) -> List[List[Dict[str, Any]]]:
    """
    Extract tables from PDF and return structured data
    """
    all_tables = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Extract tables from the page
            tables = page.extract_tables()
            
            for table in tables:
                if not table or len(table) < 2:
                    continue
                    
                # Get headers from first row
                headers = [str(cell).strip() if cell else '' for cell in table[0]]
                
                # Find relevant column indices
                sno_idx = None
                product_idx = None
                amount_excl_tax_idx = None
                discount_idx = None
                
                for i, header in enumerate(headers):
                    header_lower = header.lower()
                    if 's.no' in header_lower or 'sno' in header_lower:
                        sno_idx = i
                    elif 'product' in header_lower:
                        product_idx = i
                    elif 'amount' in header_lower and 'excl' in header_lower:
                        amount_excl_tax_idx = i
                    elif 'discount' in header_lower:
                        discount_idx = i
                
                # Skip if we don't have essential columns
                if sno_idx is None or product_idx is None or amount_excl_tax_idx is None:
                    continue
                
                # Process data rows
                table_data = []
                for row in table[1:]:  # Skip header row
                    if len(row) <= max(sno_idx, product_idx, amount_excl_tax_idx):
                        continue
                        
                    sno = str(row[sno_idx]).strip() if row[sno_idx] else ''
                    product = str(row[product_idx]).strip() if row[product_idx] else ''
                    amount_str = str(row[amount_excl_tax_idx]).strip() if row[amount_excl_tax_idx] else ''
                    discount_str = str(row[discount_idx]).strip() if discount_idx is not None and row[discount_idx] else ''
                    
                    # Skip empty rows
                    if not sno and not product and not amount_str:
                        continue
                    
                    # Parse amount (handle USD prefix, commas, negative values)
                    amount = parse_amount(amount_str)
                    
                    # Determine if there's a discount
                    has_discount = 'Y' if discount_str and discount_str != '' and discount_str != '0' else 'N'
                    
                    table_data.append({
                        'sno': sno,
                        'product': product,
                        'amount_excl_tax': amount,
                        'discount': has_discount
                    })
                
                if table_data:
                    all_tables.append(table_data)
    
    return all_tables

def parse_amount(amount_str: str) -> float:
    """
    Parse amount string to float, handling USD prefix, commas, negatives
    """
    if not amount_str or amount_str == '':
        return 0.0
    
    # Remove USD prefix, whitespace, and commas
    cleaned = re.sub(r'USD\s*', '', amount_str, flags=re.IGNORECASE)
    cleaned = cleaned.replace(',', '').strip()
    
    # Handle negative values
    is_negative = cleaned.startswith('-')
    if is_negative:
        cleaned = cleaned[1:]
    
    try:
        value = float(cleaned)
        return -value if is_negative else value
    except ValueError:
        return 0.0

def format_product_text(product_text: str) -> str:
    """
    Format product text with line breaks before Entitlement Number and Billing period
    """
    if not product_text:
        return product_text
    
    # Add line breaks before "Entitlement Number:" and "Billing period:"
    formatted_text = re.sub(r'(Entitlement Number:)', r'\n\1', product_text)
    formatted_text = re.sub(r'(Billing period:)', r'\n\1', formatted_text)
    
    return formatted_text

def group_by_sno(table_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Group items by S.no, handling cases where multiple items share the same S.no
    """
    grouped = {}
    current_sno = None
    
    for item in table_data:
        sno = item['sno']
        
        # If S.no is empty, it belongs to the previous S.no
        if not sno and current_sno:
            sno = current_sno
        else:
            current_sno = sno
        
        if sno not in grouped:
            grouped[sno] = []
        
        grouped[sno].append(item)
    
    return grouped

def create_excel_output(grouped_data: Dict[str, List[Dict[str, Any]]], output_path: str):
    """
    Create Excel file with enhanced formatting and formulas in column E
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"
    
    # Set headers
    ws['A1'] = 'S.no'
    ws['B1'] = 'Discount?'
    ws['C1'] = 'Product'
    ws['D1'] = 'Amount excl. tax'
    ws['E1'] = 'Total'
    
    # Format header row
    header_font = Font(bold=True)
    for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws[col].font = header_font
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 60  # Much wider for product descriptions
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # Create styles
    red_font = Font(color="FF0000")  # Red color for totals
    wrap_text = Alignment(wrap_text=True, vertical='top')
    currency_format = '#,##0.00'
    
    current_row = 2
    formula_rows = []  # Track rows with formulas for grand total
    
    # Sort by S.no for consistent output
    for sno in sorted(grouped_data.keys(), key=lambda x: int(x) if x.isdigit() else 999):
        items = grouped_data[sno]
        
        if not items:
            continue
        
        # First item in the group
        first_item = items[0]
        
        # Set values
        ws[f'A{current_row}'] = sno
        ws[f'B{current_row}'] = first_item['discount']
        
        # Format product text with line breaks
        formatted_product = format_product_text(first_item['product'])
        ws[f'C{current_row}'] = formatted_product
        ws[f'C{current_row}'].alignment = wrap_text
        
        # Format amount with currency
        ws[f'D{current_row}'] = first_item['amount_excl_tax']
        ws[f'D{current_row}'].number_format = currency_format
        
        # Create formula for group total
        group_start_row = current_row
        group_end_row = current_row + len(items) - 1
        
        if len(items) == 1:
            # Single item - formula just references the single cell
            formula = f"=D{current_row}"
        else:
            # Multiple items - sum the range
            formula = f"=SUM(D{group_start_row}:D{group_end_row})"
        
        ws[f'E{current_row}'] = formula
        ws[f'E{current_row}'].number_format = currency_format
        ws[f'E{current_row}'].font = red_font  # Make total red
        
        # Track this row for grand total formula
        formula_rows.append(current_row)
        
        # Set row height to accommodate wrapped text
        ws.row_dimensions[current_row].height = None  # Auto-fit
        
        current_row += 1
        
        # Additional items under the same S.no (sub-items)
        for item in items[1:]:
            ws[f'D{current_row}'] = item['amount_excl_tax']
            ws[f'D{current_row}'].number_format = currency_format
            current_row += 1
    
    # Add grand total with formula
    ws[f'D{current_row}'] = "TOTAL"
    ws[f'D{current_row}'].font = Font(bold=True)
    
    # Create grand total formula that sums all the group totals
    if formula_rows:
        formula_cells = [f"E{row}" for row in formula_rows]
        grand_total_formula = f"=SUM({','.join(formula_cells)})"
        ws[f'E{current_row}'] = grand_total_formula
    else:
        ws[f'E{current_row}'] = 0
    
    ws[f'E{current_row}'].number_format = currency_format
    ws[f'E{current_row}'].font = Font(color="FF0000", bold=True)  # Red and bold for grand total
    
    # Auto-fit row heights for better text display
    for row in range(2, current_row + 1):
        ws.row_dimensions[row].height = None
    
    # Save the workbook
    wb.save(output_path)
    print(f"Formatted Excel file with formulas saved to: {output_path}")

def process_document_to_excel(input_path: str, excel_path: str):
    """
    Main function to process PDF or JSON and create Excel output
    """
    print(f"Processing file: {input_path}")
    
    # Detect input type
    input_type = detect_input_type(input_path)
    print(f"Detected input type: {input_type.upper()}")
    
    # Extract data based on input type
    if input_type == 'json':
        table_data = extract_from_json(input_path)
        if not table_data:
            print("No data extracted from JSON")
            return
        combined_data = table_data
    else:  # PDF
        all_tables = extract_tables_from_pdf(input_path)
        if not all_tables:
            print("No tables found in PDF")
            return
        # Combine all tables
        combined_data = []
        for table in all_tables:
            combined_data.extend(table)
    
    print(f"Found {len(combined_data)} items")
    
    # Group by S.no
    grouped_data = group_by_sno(combined_data)
    
    print(f"Grouped into {len(grouped_data)} S.no entries")
    
    # Create Excel output with enhanced formatting and formulas
    create_excel_output(grouped_data, excel_path)
    
    print("Processing complete!")

# Convenience functions for specific input types
def process_pdf_to_excel(pdf_path: str, excel_path: str):
    """Process PDF specifically"""
    process_document_to_excel(pdf_path, excel_path)

def process_json_to_excel(json_path: str, excel_path: str):
    """Process JSON specifically"""
    process_document_to_excel(json_path, excel_path)

# Example usage
if __name__ == "__main__":
    # Install required packages first:
    # pip install pdfplumber pandas openpyxl
    
    input_file = "your_document.pdf"  # Can be .pdf or .json
    # OR
    # input_file = "your_data.json"
    
    excel_file = "extracted_data.xlsx"  # Output Excel file
    
    process_document_to_excel(input_file, excel_file)
    
    # You can also use specific functions if you know the input type:
    # process_pdf_to_excel("document.pdf", "output.xlsx")
    # process_json_to_excel("data.json", "output.xlsx")

if __name__ == "__main__":

    import os
    pdf_files = [f for f in os.listdir('quotes') if f.lower().endswith('.pdf')]
    for pdf_file in pdf_files:
        excel_file = "output/" + os.path.splitext(pdf_file)[0] + '_extracted.xlsx'
        pdf_file = os.path.join('quotes/', pdf_file)
        process_pdf_to_excel(pdf_file, excel_file)